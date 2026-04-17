import openpyxl
from collections import Counter

def check_lead_names(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    try:
        idx = headers.index('Lead Name')
    except ValueError:
        print("Error: 'Lead Name' column not found.")
        return

    names = []
    original_row_mapping = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        val = row[idx].value
        name = str(val).strip() if val is not None else ""
        names.append(name)
        if name not in original_row_mapping:
            original_row_mapping[name] = []
        original_row_mapping[name].append(row_idx)

    total = len(names)
    empty = [i+2 for i, n in enumerate(names) if not n or n.lower() == 'none' or n == '0']
    
    counts = Counter(names)
    duplicates = {n: original_row_mapping[n] for n, count in counts.items() if count > 1 and n != ""}

    print(f"Total Leads: {total}")
    print(f"Empty or Invalid Names: {len(empty)}")
    if empty:
        print(f"  Rows: {empty[:20]}...")
        
    print(f"Unique Names: {len(counts) - (1 if '' in counts else 0)}")
    print(f"Duplicate Name Groups: {len(duplicates)}")
    
    if duplicates:
        print("\n--- Top 10 Duplicate Groups ---")
        for i, (name, rows) in enumerate(list(duplicates.items())[:10]):
            print(f"{i+1}. \"{name}\": occurring in rows {rows}")

if __name__ == "__main__":
    check_lead_names('Zam Leads to map with jason file.xlsx')
