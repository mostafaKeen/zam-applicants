import openpyxl

def analyze_fields(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1] if cell.value]
    total_rows = ws.max_row - 1
    
    results = []
    
    for col_idx, header in enumerate(headers, start=1):
        values = []
        empty_count = 0
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            val = row[0].value
            if val is None or str(val).strip() == "" or str(val).lower() == 'none':
                empty_count += 1
            else:
                values.append(str(val).strip())
        
        unique_values = len(set(values))
        results.append({
            'header': header,
            'unique': unique_values,
            'empty': empty_count,
            'coverage': ((total_rows - empty_count) / total_rows) * 100 if total_rows > 0 else 0
        })

    # Sort candidates by coverage (least empty first) and then uniqueness
    # We want fields with 0 empty values and highest uniqueness
    candidates = sorted(results, key=lambda x: (x['empty'], -x['unique']))

    print(f"Total Rows: {total_rows}\n")
    print(f"{'Field Name':<35} | {'Unique':<8} | {'Empty':<8} | {'Coverage':<8}")
    print("-" * 75)
    for res in candidates[:20]: # Show top 20 candidates
        print(f"{res['header']:<35} | {res['unique']:<8} | {res['empty']:<8} | {res['coverage']:.1f}%")

if __name__ == "__main__":
    analyze_fields('Zam Leads to map with jason file.xlsx')
