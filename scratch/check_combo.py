import openpyxl
from collections import Counter

def check_combination(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    name_idx = headers.index('Lead Name')
    mobile_idx = headers.index('Mobile')
    work_phone_idx = headers.index('Work Phone')
    
    combinations = []
    for row in ws.iter_rows(min_row=2):
        name = str(row[name_idx].value).strip() if row[name_idx].value else ""
        mobile = str(row[mobile_idx].value).strip() if row[mobile_idx].value else ""
        work = str(row[work_phone_idx].value).strip() if row[work_phone_idx].value else ""
        
        phone = mobile if mobile else work
        combinations.append(f"{name}|{phone}")

    counts = Counter(combinations)
    total = len(combinations)
    unique = len(counts)
    duplicates = {c: count for c, count in counts.items() if count > 1}
    
    print(f"Total rows: {total}")
    print(f"Unique Name+Phone combos: {unique}")
    print(f"Duplicate combos: {len(duplicates)}")
    
    if duplicates:
        print("\nExamples of duplicate Name+Phone:")
        for combo, count in list(duplicates.items())[:10]:
            print(f"  - {combo} (count: {count})")

if __name__ == "__main__":
    check_combination('Zam Leads to map with jason file.xlsx')
