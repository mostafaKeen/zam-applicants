import openpyxl
from collections import Counter

def check_name_day_combo(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    name_idx = headers.index('Lead Name')
    created_idx = headers.index('Created')
    
    combinations = []
    for row in ws.iter_rows(min_row=2):
        name = str(row[name_idx].value).strip() if row[name_idx].value else ""
        created = str(row[created_idx].value).strip() if row[created_idx].value else ""
        # Extract only YYYY-MM-DD
        day = created.split(' ')[0] if ' ' in created else created
        combinations.append(f"{name}|{day}")

    counts = Counter(combinations)
    total = len(combinations)
    unique = len(counts)
    duplicates = {c: count for c, count in counts.items() if count > 1}
    
    print(f"Total rows: {total}")
    print(f"Unique Name+Day combos: {unique}")
    print(f"Duplicate combos: {len(duplicates)}")
    
    if duplicates:
        print("\nExamples of duplicate Name+Day:")
        for combo, count in list(duplicates.items())[:10]:
            print(f"  - {combo} (count: {count})")

if __name__ == "__main__":
    check_name_day_combo('Zam Leads to map with jason file.xlsx')
