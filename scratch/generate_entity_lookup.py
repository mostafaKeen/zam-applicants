import openpyxl
import json

def generate_entity_lookup(excel_path, output_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    id_idx = headers.index('EntityId')
    name_idx = headers.index('Lead Name')
    created_idx = headers.index('Created')
    
    # EntityId -> {Name, Created}
    mapping = {}
    for row in ws.iter_rows(min_row=2):
        eid = str(row[id_idx].value).strip() if row[id_idx].value else None
        if not eid: continue
        
        name = str(row[name_idx].value).strip() if row[name_idx].value else ""
        # The date needs to be consistent for matching. 
        # Excel gives datetime object or string.
        created = str(row[created_idx].value).strip() if row[created_idx].value else ""
        
        mapping[eid] = {
            "name": name,
            "created": created
        }

    with open(output_path, 'w') as f:
        json.dump(mapping, f, indent=4)
    print(f"Generated lookup for {len(mapping)} entities in {output_path}")

if __name__ == "__main__":
    generate_entity_lookup('Zam Leads to map with jason file.xlsx', 'entity_to_name_date.json')
