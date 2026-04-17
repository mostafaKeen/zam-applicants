import openpyxl
import json
import re

def clean_phone(phone):
    if phone is None:
        return None
    phone = str(phone).strip()
    if not phone:
        return None
    # Remove non-numeric characters except for leading +
    # But usually Bitrix stores phone numbers as digits or with +
    # We want to keep it simple for matching
    cleaned = re.sub(r'[^\d+]', '', phone)
    return cleaned

def extract_mapping(excel_path, output_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    
    try:
        entity_id_idx = headers.index('EntityId')
        work_phone_idx = headers.index('Work Phone')
        mobile_idx = headers.index('Mobile')
    except ValueError as e:
        print(f"Error: Could not find required headers. {e}")
        return

    mapping = {}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        entity_id = row[entity_id_idx].value
        work_phone = row[work_phone_idx].value
        mobile = row[mobile_idx].value
        
        if entity_id is None:
            continue
            
        entity_id_str = str(entity_id).strip()
        
        # Prefer mobile, then work phone
        phone = mobile if mobile else work_phone
        cleaned_phone = clean_phone(phone)
        
        if cleaned_phone:
            mapping[entity_id_str] = cleaned_phone

    with open(output_path, 'w') as f:
        json.dump(mapping, f, indent=4)
    
    print(f"Extracted {len(mapping)} mapping entries to {output_path}")

if __name__ == "__main__":
    extract_mapping('Zam Leads to map with jason file.xlsx', 'lead_mapping.json')
